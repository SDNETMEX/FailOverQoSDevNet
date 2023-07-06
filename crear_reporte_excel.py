from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook


def integer_to_excel_column(col_number):
    loops_number = 0
    while col_number > 90:
        loops_number = loops_number + 1
        col_number = col_number - 26

    if loops_number > 0:
        return chr(64 + loops_number) + chr(col_number)
    else:
        return chr(col_number)


def crear_reporte_excel(data, nombre_archivo):
    wb = Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(sheet)
    ws = wb.create_sheet(0)
    ws.title = "Reporte"
    #font = Font(color="00FF0000")

    int_ascii = 65
    for x in range(len(data[0])):
        columna = integer_to_excel_column(int_ascii)
        if int_ascii == 65:
            ws.column_dimensions[columna].width = 20
        elif int_ascii == 67:
            ws.column_dimensions[columna].width = 30
        elif int_ascii == 68:
            ws.column_dimensions[columna].width = 50
        elif int_ascii == 72:
            ws.column_dimensions[columna].width = 20
        elif int_ascii == 73:
            ws.column_dimensions[columna].width = 30
        elif int_ascii == 74 or int_ascii == 75:
            ws.column_dimensions[columna].width = 15
        else:
            ws.column_dimensions[columna].width = 12
        ws[columna + '1'].fill = PatternFill("solid", fgColor="00FFFF00")
        ws[columna + '1'] = data[0][x]
        int_ascii += 1
    cont_row = 1
    for row in data[1:]:
        int_ascii = 65
        for x in range(len(row)):
            columna = integer_to_excel_column(int_ascii)
            ws[columna + str(cont_row + 1)] = str(data[cont_row][x])
            int_ascii = int_ascii + 1
        cont_row = cont_row + 1

    wb.save("/var/www/FOyQoS/mail_excel/" + nombre_archivo)
    return 200